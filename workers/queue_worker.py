import json
import os
import time
import shutil
import zipfile
from pathlib import Path
from datetime import datetime, timezone, timedelta

ROOT = Path("/home/frank/projects/TWICKELL_ROOT")


REQ_DIR = ROOT / "bridge_queue" / "requests"
RES_DIR = ROOT / "bridge_queue" / "results"
JOBS_DIR = ROOT / "jobs"
UPLOADS_DIR = ROOT / "uploads"
OUTPUTS_DIR = ROOT / "outputs"
LOG_PATH = ROOT / "logs" / "queue_worker.log"

POLL_SECONDS = 1.0
CLEANUP_EVERY_SECONDS = 60
RETENTION_ZERO_DONE_BUFFER_SECONDS = 3600
def utcnow_iso():
    return datetime.now(timezone.utc).isoformat()

def parse_iso(s):
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)

def log(msg):
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"{utcnow_iso()}  {msg}\n")
    print(msg, flush=True)

def read_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    os.replace(tmp, path)

def job_path(job_id):
    return JOBS_DIR / f"{job_id}.json"

def output_dir(job_id):
    return OUTPUTS_DIR / job_id

def output_xlsx_path(job_id):
    return output_dir(job_id) / "output.xlsx"

def bundle_zip_path(job_id):
    return output_dir(job_id) / "bundle.zip"

def build_zip_bundle(job):
    jid = job.get("job_id") or job.get("id")
    if not jid:
        raise RuntimeError("job missing id")
    out_dir = output_dir(jid)
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx = Path(job["paths"]["output_xlsx"])
    meta_path = out_dir / "metadata.json"
    summary_path = out_dir / "summary.txt"
    zip_path = bundle_zip_path(jid)

    metadata = {
        "job_id": jid,
        "email": job.get("email"),
        "created_at": job.get("created_at"),
        "done_at": job.get("done_at"),
        "retention_days": job.get("retention_days"),
        "pages": job.get("pages"),
        "amount_due": job.get("amount_due"),
        "free_sample_applied": job.get("free_sample_applied"),
    }
    write_json(meta_path, metadata)

    summary_lines = [
        f"job_id: {jid}",
        f"status: {job.get('status')}",
        f"pages: {job.get('pages')}",
        f"amount_due: {job.get('amount_due')}",
        f"free_sample_applied: {job.get('free_sample_applied')}",
        f"retention_days: {job.get('retention_days')}",
        f"created_at: {job.get('created_at')}",
        f"done_at: {job.get('done_at')}",
    ]
    summary_path.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        z.write(xlsx, arcname="output.xlsx")
        z.write(meta_path, arcname="metadata.json")
        z.write(summary_path, arcname="summary.txt")

        # Include original PDF if present
        pdf_p = Path(job.get("paths", {}).get("input_pdf", ""))
        if pdf_p.exists() and pdf_p.is_file():
            z.write(pdf_p, arcname="input.pdf")

    return zip_path

def generate_xlsx_for_job(job):
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("openpyxl is required (pip install openpyxl).") from e

    jid = job.get("job_id") or job.get("id")
    if not jid:
        raise RuntimeError("job missing id")
    out_dir = output_dir(jid)
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_xlsx_path(jid)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Twickell"
    ws["A2"] = "job_id"
    ws["B2"] = jid
    ws["A3"] = "pages"
    ws["B3"] = job.get("pages")
    ws["A4"] = "email"
    ws["B4"] = job.get("email")
    ws["A5"] = "note"
    ws["B5"] = "Replace this generator with real PDF→XLSX extraction."

    wb.save(xlsx_path)
    return xlsx_path

def process_request(req_path):
    req = read_json(req_path)
    job_id = req.get("job_id")
    if not job_id:
        log(f"SKIP bad request (no job_id): {req_path.name}")
        req_path.unlink(missing_ok=True)
        return

    jp = job_path(job_id)
    if not jp.exists():
        log(f"JOB_NOT_FOUND {job_id} (request {req_path.name})")
        req_path.unlink(missing_ok=True)
        return

    job = read_json(jp)
    log(f"PROCESSING {req_path.name}")

    xlsx_path = generate_xlsx_for_job(job)

    job.setdefault("paths", {})
    job["paths"]["output_xlsx"] = str(xlsx_path)
    job["status"] = "done"
    job["done_at"] = utcnow_iso()

    zip_path = build_zip_bundle(job)
    job["paths"]["bundle_zip"] = str(zip_path)

    write_json(jp, job)

    res = {"job_id": job_id, "ok": True, "done_at": job["done_at"], "paths": job["paths"]}
    write_json(RES_DIR / f"{job_id}.json", res)

    req_path.unlink(missing_ok=True)
    log(f"DONE {req_path.name}")

def should_delete(job):
    if job.get("status") not in ("done", "failed"):
        return False

    created_at = job.get("created_at")
    if not created_at:
        return False

    retention_days = int(job.get("retention_days", 7))
    now = datetime.now(timezone.utc)

    if retention_days == 0:
        done_at = job.get("done_at")
        if not done_at:
            return False
        return now >= (parse_iso(done_at) + timedelta(seconds=RETENTION_ZERO_DONE_BUFFER_SECONDS))

    return now >= (parse_iso(created_at) + timedelta(days=retention_days))


def delete_job_everywhere(job_id):
    targets = [
        UPLOADS_DIR / job_id,
        OUTPUTS_DIR / job_id,
        RES_DIR / f"{job_id}.json",
        REQ_DIR / f"{job_id}.json",
        JOBS_DIR / f"{job_id}.json",
    ]

    for t in targets:
        try:
            if t.is_dir():
                shutil.rmtree(t, ignore_errors=True)
            else:
                t.unlink(missing_ok=True)
        except Exception:
            pass

    log(f"CLEANUP_DELETED {job_id}")


def cleanup_pass():
    if not JOBS_DIR.exists():
        return
    for jp in JOBS_DIR.glob("*.json"):
        try:
            job = read_json(jp)
            job_id = job.get("job_id") or jp.stem
            if should_delete(job):
                delete_job_everywhere(job_id)
        except Exception:
            continue

def main():
    REQ_DIR.mkdir(parents=True, exist_ok=True)
    log("QUEUE_WORKER_START")

    last_cleanup = 0.0

    while True:
        now = time.time()
        if now - last_cleanup >= CLEANUP_EVERY_SECONDS:
            cleanup_pass()
            last_cleanup = now

        reqs = sorted(REQ_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime)
        if not reqs:
            time.sleep(POLL_SECONDS)
            continue

        for req_path in reqs:
            try:
                process_request(req_path)
            except Exception as e:
                log(f"ERROR processing {req_path.name}: {e}")
                req_path.unlink(missing_ok=True)


if __name__ == "__main__":
    main()

